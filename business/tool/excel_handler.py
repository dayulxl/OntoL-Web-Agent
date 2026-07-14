"""
通用 Excel 读写工具 — 无业务逻辑，可跨域引用。

使用方式:
    from business.tool.excel_handler import write_excel, read_excel, excel_response

    # 写入
    cols = [{"key":"code","label":"编码","width":15}, {"key":"name","label":"名称","width":20}]
    rows = [{"code":"X1","name":"字段一"}, {"code":"X2","name":"字段二"}]
    write_excel("/tmp/out.xlsx", "Sheet1", cols, rows)

    # 读取
    headers, rows = read_excel("/tmp/in.xlsx", "Sheet1")
    # rows = [{"code":"X1","name":"字段一"}, ...]

    # FastAPI 下载
    return excel_response("/tmp/out.xlsx", "模板.xlsx")
"""
import io
from pathlib import Path
from typing import Optional
from urllib.parse import quote

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse


# ── 样式常量 ──
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
ACTION_FONT = Font(name="微软雅黑", size=10, color="0070C0")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_excel(
    filepath: str,
    sheet_name: str,
    columns: list[dict],
    rows: list[dict],
    *,
    freeze_row: int = 2,
    validations: list[dict] | None = None,
) -> str:
    """
    写入 Excel 文件。

    columns: [{"key": "code", "label": "编码", "width": 15}, ...]
    rows:    [{"code": "X1", "name": "字段一"}, ...]
    validations: 可选，下拉校验列表，格式:
        [{"col": 1, "options": ["新增","修改","删除"], "prompt": "请选择", "allow_blank": True}, ...]
        col 为 1-based 列号；options 为下拉可选项；prompt 为输入提示；allow_blank 允许空值。
    返回 filepath
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 表头
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col["label"])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = col.get("width", 15)

    # 数据行
    last_data_row = len(rows) + 1  # 1-based, last row with data
    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(columns, 1):
            val = row.get(col["key"], "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.alignment = CENTER if ci <= 2 else LEFT
            cell.border = THIN_BORDER

    # ── 数据校验（下拉列表） ──
    if validations:
        for vd in validations:
            col_letter = get_column_letter(vd["col"])
            options_str = ",".join(vd["options"])
            # 下拉范围：从第 2 行到最后一行数据 + 额外空白行
            end_row = max(last_data_row, freeze_row + 1)
            cell_range = f"{col_letter}2:{col_letter}{end_row}"
            dv = DataValidation(
                type="list",
                formula1=f'"{options_str}"',
                allow_blank=vd.get("allow_blank", True),
            )
            dv.error = "请从下拉列表中选择"
            dv.errorTitle = "输入无效"
            if vd.get("prompt"):
                dv.prompt = vd["prompt"]
                dv.promptTitle = "操作提示"
            ws.add_data_validation(dv)
            dv.add(cell_range)

    # 冻结
    ws.freeze_panes = f"A{freeze_row}"
    # 自动筛选
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{last_data_row}"

    wb.save(filepath)
    return filepath


def read_excel(filepath: str, sheet_name: Optional[str] = None) -> tuple[list[str], list[dict]]:
    """
    读取 Excel 文件，返回 (headers, rows)。

    headers: ["编码", "名称", ...]
    rows:    [{"编码": "X1", "名称": "字段一"}, ...]
    列名用表头原始值作为 key
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    headers_raw = [str(h).strip() if h else "" for h in next(rows_iter)]
    # 去重：同名列加序号
    seen: dict[str, int] = {}
    headers: list[str] = []
    for h in headers_raw:
        if h in seen:
            seen[h] += 1
            headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            headers.append(h)

    rows: list[dict] = []
    for row in rows_iter:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        d = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else ""
            d[h] = str(val).strip() if val is not None else ""
        rows.append(d)

    wb.close()
    return headers, rows


def excel_response(filepath: str, download_name: str) -> StreamingResponse:
    """FastAPI StreamingResponse，用于文件下载。"""
    data = Path(filepath).read_bytes()
    encoded_name = quote(download_name.encode("utf-8", errors="replace"))
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"},
    )
